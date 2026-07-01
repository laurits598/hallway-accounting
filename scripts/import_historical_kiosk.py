#!/usr/bin/env python3
"""Import monthly kiosk aggregate XLSX exports into the purchases table."""

import argparse
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = PROJECT_ROOT / "kollegianeren.db"
DEFAULT_INPUT_DIR = PROJECT_ROOT / "data" / "historical_kiosk"
HISTORICAL_PRODUCT = "Historical kiosk total"
FILENAME_RE = re.compile(r"^(\d{1,2})_\d{1,2}_(\d{4})_\d{1,2}_\d{1,2}_\d{4}\.xlsx$")


def month_bounds(month, year):
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year + 1:04d}-01-01" if month == 12 else f"{year:04d}-{month + 1:02d}-01"
    return start, end


def read_export(path):
    match = FILENAME_RE.match(path.name)
    if not match:
        raise ValueError(f"Unsupported historical kiosk filename: {path.name}")
    month, year = map(int, match.groups())

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header = [str(value or "").strip().casefold() for value in next(worksheet.iter_rows(values_only=True))]
        if header[:3] != ["name", "room", "total"]:
            raise ValueError(f"Unexpected columns in {path.name}; expected name, room, total")

        totals = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None for value in row):
                continue
            name = str(row[0] or "").strip() or f"Historical resident {row_number}"
            room = str(row[1] or "").strip()
            if not room:
                raise ValueError(f"Missing room in {path.name}, row {row_number}")
            try:
                total = round(float(row[2]), 2)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Invalid total in {path.name}, row {row_number}") from exc
            totals.append((name, room, total))
    finally:
        workbook.close()
    return month, year, totals


def import_export(connection, path):
    month, year, totals = read_export(path)
    start, end = month_bounds(month, year)
    existing = connection.execute(
        "SELECT COUNT(*) FROM purchases WHERE timestamp >= ? AND timestamp < ?",
        (start, end),
    ).fetchone()[0]
    if existing:
        return {"status": "skipped", "month": month, "year": year, "rows": existing}

    connection.execute(
        """
        INSERT OR IGNORE INTO products (name, retail_price, price, image, active)
        VALUES (?, NULL, 0, NULL, 0)
        """,
        (HISTORICAL_PRODUCT,),
    )
    product_id = connection.execute(
        "SELECT id FROM products WHERE name = ?", (HISTORICAL_PRODUCT,)
    ).fetchone()[0]

    timestamp = f"{year:04d}-{month:02d}-01T12:00:00"
    inserted = 0
    for name, room, total in totals:
        resident = connection.execute(
            "SELECT rid FROM residents WHERE room = ? ORDER BY active DESC, rid DESC LIMIT 1",
            (room,),
        ).fetchone()
        if resident is None:
            historical_name = name
            if connection.execute(
                "SELECT 1 FROM residents WHERE name = ?", (historical_name,)
            ).fetchone():
                historical_name = f"{name} (historical room {room})"
            cursor = connection.execute(
                "INSERT INTO residents (name, room, active) VALUES (?, ?, 0)",
                (historical_name, room),
            )
            resident = (cursor.lastrowid,)
        connection.execute(
            """
            INSERT INTO purchases (product_id, resident_id, quantity, price, timestamp)
            VALUES (?, ?, 1, ?, ?)
            """,
            (product_id, resident[0], total, timestamp),
        )
        inserted += 1

    return {"status": "imported", "month": month, "year": year, "rows": inserted}


def import_directory(db_path=DEFAULT_DB, input_dir=DEFAULT_INPUT_DIR):
    paths = sorted(Path(input_dir).glob("*.xlsx"))
    if not paths:
        print(f"No historical kiosk exports found in {input_dir}")
        return []

    connection = sqlite3.connect(db_path)
    results = []
    try:
        for path in paths:
            with connection:
                result = import_export(connection, path)
            results.append(result)
            print(
                f"{path.name}: {result['status']} "
                f"({result['rows']} rows, {result['month']:02d}/{result['year']})"
            )
    finally:
        connection.close()
    return results


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    args = parser.parse_args()
    import_directory(args.db, args.input_dir)


if __name__ == "__main__":
    main()
