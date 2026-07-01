#!/usr/bin/env python3
"""
Import historical accounting sheets from data/historical/ into SQLite.

Each XLSX file contains one month's accounting matrix:
- columns 1-3: resident name, room, monthly total
- remaining columns: product names with purchase counts

This importer:
- infers per-product unit prices from each sheet using least-squares
- creates missing residents/products as inactive catalog entries
- inserts one aggregated purchase row per resident/product/month
- uses deterministic timestamps so rerunning it does not duplicate imports
"""

from __future__ import annotations

import sqlite3
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DB_PATH = PROJECT_ROOT / "kollegianeren.db"
HISTORICAL_DATA_DIR = PROJECT_ROOT / "data" / "historical"


@dataclass
class SheetRow:
    resident_name: str
    room: str
    total: float
    counts: dict[str, int]


@dataclass
class SheetData:
    path: Path
    month_end: datetime
    headers: list[str]
    rows: list[SheetRow]
    inferred_prices: dict[str, float]
    max_error: float


def parse_month_end(path: Path) -> datetime:
    parts = path.stem.split("_")
    if len(parts) != 6:
        raise ValueError(f"Unexpected filename format: {path.name}")
    month, day, year = map(int, parts[3:6])
    return datetime(year, month, day, 12, 0, 0)


def normalize_product_name(name: str) -> str:
    return str(name or "").strip()


def load_sheet(path: Path) -> SheetData:
    ws = load_workbook(path, data_only=True).active
    values = list(ws.iter_rows(values_only=True))
    headers = [normalize_product_name(value) for value in values[0][3:] if value is not None]

    rows: list[SheetRow] = []
    matrix: list[list[float]] = []
    totals: list[float] = []

    for raw_row in values[1:]:
        name = str(raw_row[0] or "").strip()
        if not name:
            continue
        room = str(raw_row[1] or "").strip()
        total = float(raw_row[2] or 0)
        count_values = [int(value or 0) for value in raw_row[3:3 + len(headers)]]
        counts = {header: count for header, count in zip(headers, count_values) if count > 0}
        rows.append(SheetRow(resident_name=name, room=room, total=total, counts=counts))
        matrix.append([float(count) for count in count_values])
        totals.append(total)

    coefficients = np.array(matrix, dtype=float)
    targets = np.array(totals, dtype=float)
    solution, _, _, _ = np.linalg.lstsq(coefficients, targets, rcond=None)
    inferred_prices = {header: float(price) for header, price in zip(headers, solution)}
    max_error = float(np.max(np.abs(coefficients @ solution - targets)))

    return SheetData(
        path=path,
        month_end=parse_month_end(path),
        headers=headers,
        rows=rows,
        inferred_prices=inferred_prices,
        max_error=max_error,
    )


def representative_prices(sheets: Iterable[SheetData]) -> dict[str, float]:
    grouped: dict[str, list[float]] = {}
    for sheet in sheets:
        if sheet.max_error > 0.01:
            continue
        for name, price in sheet.inferred_prices.items():
            grouped.setdefault(name, []).append(round(price, 2))

    selected: dict[str, float] = {}
    for name, prices in grouped.items():
        most_common = Counter(prices).most_common(1)[0][0]
        selected[name] = float(most_common)
    return selected


def ensure_resident(cursor: sqlite3.Cursor, name: str, room: str) -> int:
    cursor.execute("SELECT rid FROM residents WHERE name = ?", (name,))
    row = cursor.fetchone()
    if row:
        resident_id = int(row[0])
        if room:
            cursor.execute("UPDATE residents SET room = COALESCE(NULLIF(room, ''), ?), updated_at = CURRENT_TIMESTAMP WHERE rid = ?", (room, resident_id))
        return resident_id

    cursor.execute(
        """
        INSERT INTO residents (name, room, image, active)
        VALUES (?, ?, NULL, 0)
        """,
        (name, room or "unknown"),
    )
    return int(cursor.lastrowid)


def ensure_product(cursor: sqlite3.Cursor, name: str, fallback_price: float) -> int:
    cursor.execute("SELECT id FROM products WHERE name = ?", (name,))
    row = cursor.fetchone()
    if row:
        return int(row[0])

    cursor.execute(
        """
        INSERT INTO products (name, retail_price, price, image, active)
        VALUES (?, NULL, ?, NULL, 0)
        """,
        (name, fallback_price),
    )
    return int(cursor.lastrowid)


def purchase_exists(
    cursor: sqlite3.Cursor,
    product_id: int,
    resident_id: int,
    quantity: int,
    price: float,
    timestamp: str,
) -> bool:
    cursor.execute(
        """
        SELECT 1
        FROM purchases
        WHERE product_id = ?
          AND resident_id = ?
          AND quantity = ?
          AND ABS(price - ?) < 0.0001
          AND timestamp = ?
        LIMIT 1
        """,
        (product_id, resident_id, quantity, price, timestamp),
    )
    return cursor.fetchone() is not None


def import_sheets() -> None:
    sheet_paths = sorted(HISTORICAL_DATA_DIR.glob("*.xlsx"))
    if not sheet_paths:
        raise SystemExit("No .xlsx files found in data/historical/")

    sheets = [load_sheet(path) for path in sheet_paths]
    fallback_prices = representative_prices(sheets)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    inserted = 0
    skipped = 0
    created_products = 0
    created_residents = 0

    existing_product_names = {row[0] for row in cursor.execute("SELECT name FROM products")}
    existing_resident_names = {row[0] for row in cursor.execute("SELECT name FROM residents")}

    for sheet in sheets:
        for row_index, row in enumerate(sheet.rows):
            resident_id = ensure_resident(cursor, row.resident_name, row.room)
            if row.resident_name not in existing_resident_names:
                existing_resident_names.add(row.resident_name)
                created_residents += 1

            for column_index, product_name in enumerate(sheet.headers):
                quantity = row.counts.get(product_name, 0)
                if quantity <= 0:
                    continue

                inferred_price = round(
                    fallback_prices.get(product_name, sheet.inferred_prices[product_name]),
                    2,
                )
                product_id = ensure_product(cursor, product_name, inferred_price)
                if product_name not in existing_product_names:
                    existing_product_names.add(product_name)
                    created_products += 1

                timestamp = (sheet.month_end + timedelta(seconds=(row_index * 100) + column_index)).isoformat()
                if purchase_exists(cursor, product_id, resident_id, quantity, inferred_price, timestamp):
                    skipped += 1
                    continue

                cursor.execute(
                    """
                    INSERT INTO purchases (product_id, resident_id, quantity, price, timestamp)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (product_id, resident_id, quantity, inferred_price, timestamp),
                )
                inserted += 1

    conn.commit()
    conn.close()

    print(f"Imported sheets: {len(sheets)}")
    print(f"Inserted purchases: {inserted}")
    print(f"Skipped existing purchases: {skipped}")
    print(f"Created products: {created_products}")
    print(f"Created residents: {created_residents}")
    for sheet in sheets:
        print(f"{sheet.path.name}: max price-fit error {sheet.max_error:.4f}")


if __name__ == "__main__":
    import_sheets()
